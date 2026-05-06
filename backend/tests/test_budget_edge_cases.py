"""Edge case tests for budget feature.

Covers boundary values, malformed inputs, decimal precision, multi-account
multi-fund scenarios, and recovery from various failure modes.
"""
from __future__ import annotations
import io
from datetime import date, datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook

from backend.models import (
    Account, AccountingContext, BudgetCheck, BudgetMonth, BudgetPlan, BudgetStatus,
    DenominationType, DraftAllocations, DraftLineAllocation, Fund, FundCategory,
    Posting, RestrictionClass,
)
from backend.tools.budget_comparator import compare_to_budget
from backend.tools.spreadsheet_parser import parse_spreadsheet


# ---------- Helpers ----------

def _ctx_multi(budgets: dict, ytd: dict | None = None,
               threshold: float = 0.80) -> AccountingContext:
    """Build a context with multiple accounts/funds and budget."""
    accounts: list[Account] = []
    for acct_no in budgets.keys():
        accounts.append(Account(
            account_number=acct_no,
            account_name=f"Account {acct_no}",
            account_type="Expense", fund_id="GEN",
            restriction_class=RestrictionClass.WITHOUT_RESTRICTION,
        ))
    bp = BudgetPlan(
        fiscal_year=2026, plan_date=date(2026, 1, 1),
        accounts={k: BudgetMonth(annual_total=Decimal(v)) for k, v in budgets.items()},
        uploaded_at=datetime.utcnow(),
    )
    return AccountingContext(
        church_id="ec", church_name="Edge",
        denomination_type=DenominationType.OTHER,
        fiscal_year=2026, fiscal_year_start=date(2026, 1, 1),
        accounts=accounts,
        funds=[Fund(fund_id="GEN", fund_name="General",
                    restriction_class=RestrictionClass.WITHOUT_RESTRICTION,
                    fund_category=FundCategory.GENERAL_OPERATING)],
        budget=bp,
        ytd_actuals={k: Decimal(v) for k, v in (ytd or {}).items()},
        budget_warning_threshold=threshold,
    )


def _draft_multi(postings: list[tuple[str, str, str]]) -> DraftAllocations:
    """postings: list of (line_id, account_number, debit_amount)."""
    by_line: dict[str, list] = {}
    for lid, acct, amt in postings:
        by_line.setdefault(lid, []).append(Posting(
            account_number=acct, account_name=f"Account {acct}",
            fund_id="GEN", fund_name="General",
            debit_amount=Decimal(amt),
            restriction_class=RestrictionClass.WITHOUT_RESTRICTION,
        ))
    lines = [
        DraftLineAllocation(
            line_id=lid, description=f"Line {lid}",
            postings=postings_for_line,
            total_debits=sum((p.debit_amount for p in postings_for_line), Decimal("0")),
            total_credits=Decimal("0"),
            balanced=False,
        )
        for lid, postings_for_line in by_line.items()
    ]
    total = sum((l.total_debits for l in lines), Decimal("0"))
    return DraftAllocations(
        invoice_number="INV", lines=lines,
        document_total_debits=total, document_total_credits=Decimal("0"),
        document_balanced=False,
    )


# ---------- Boundary value tests ----------

def test_exactly_at_warning_threshold_is_within():
    """80.0% exactly should NOT trigger warning (uses strict >)."""
    ctx = _ctx_multi({"7100": "100"}, ytd={"7100": "0"}, threshold=0.80)
    draft = _draft_multi([("L1", "7100", "80")])
    results = compare_to_budget(draft, ctx)
    # 80/100 = 80% — strictly > 0.80 is false, so WITHIN
    assert results[0].status == BudgetStatus.WITHIN_BUDGET


def test_just_above_warning_threshold():
    ctx = _ctx_multi({"7100": "100"}, ytd={"7100": "0"}, threshold=0.80)
    draft = _draft_multi([("L1", "7100", "80.01")])
    results = compare_to_budget(draft, ctx)
    assert results[0].status == BudgetStatus.WARNING


def test_exactly_at_annual_budget_is_within():
    """100/100 = 100% is OVER (after > annual fails when after == annual)."""
    ctx = _ctx_multi({"7100": "100"}, ytd={"7100": "0"})
    draft = _draft_multi([("L1", "7100", "100")])
    results = compare_to_budget(draft, ctx)
    # after == annual; pct == 1.0; not > annual so WARNING (above 0.80)
    assert results[0].status == BudgetStatus.WARNING


def test_one_cent_over():
    ctx = _ctx_multi({"7100": "100"}, ytd={"7100": "99.99"})
    draft = _draft_multi([("L1", "7100", "0.02")])
    results = compare_to_budget(draft, ctx)
    assert results[0].status == BudgetStatus.OVER_BUDGET


def test_threshold_zero_falls_back_to_default():
    """budget_warning_threshold=0 is falsy → comparator falls back to 0.80."""
    ctx = _ctx_multi({"7100": "100"}, threshold=0.0)
    # 1/100 = 1% — under default 80% threshold → WITHIN
    draft = _draft_multi([("L1", "7100", "1")])
    results = compare_to_budget(draft, ctx)
    assert results[0].status == BudgetStatus.WITHIN_BUDGET


def test_threshold_one_means_only_over_triggers():
    ctx = _ctx_multi({"7100": "100"}, threshold=1.0)
    draft = _draft_multi([("L1", "7100", "99.99")])
    results = compare_to_budget(draft, ctx)
    assert results[0].status == BudgetStatus.WITHIN_BUDGET


# ---------- Empty/missing data ----------

def test_empty_draft_returns_empty():
    ctx = _ctx_multi({"7100": "1000"})
    draft = DraftAllocations(
        invoice_number="INV", lines=[],
        document_total_debits=Decimal("0"),
        document_total_credits=Decimal("0"),
        document_balanced=True,
    )
    assert compare_to_budget(draft, ctx) == []


def test_account_no_data_no_budget():
    """Account in COA but not in budget plan and no YTD."""
    bp = BudgetPlan(
        fiscal_year=2026, plan_date=date(2026, 1, 1),
        accounts={},  # empty
        uploaded_at=datetime.utcnow(),
    )
    ctx = AccountingContext(
        church_id="t", church_name="T",
        denomination_type=DenominationType.OTHER,
        fiscal_year=2026, fiscal_year_start=date(2026, 1, 1),
        accounts=[Account(account_number="7100", account_name="O",
                          account_type="Expense", fund_id="GEN",
                          restriction_class=RestrictionClass.WITHOUT_RESTRICTION)],
        funds=[Fund(fund_id="GEN", fund_name="G",
                    restriction_class=RestrictionClass.WITHOUT_RESTRICTION,
                    fund_category=FundCategory.GENERAL_OPERATING)],
        budget=bp,
    )
    draft = _draft_multi([("L1", "7100", "100")])
    results = compare_to_budget(draft, ctx)
    assert results[0].status == BudgetStatus.NO_BUDGET


def test_zero_debit_skipped():
    ctx = _ctx_multi({"7100": "100"})
    draft = _draft_multi([("L1", "7100", "0")])
    assert compare_to_budget(draft, ctx) == []


def test_negative_debit_skipped():
    """Per the comparator: this_invoice <= 0 is skipped."""
    ctx = _ctx_multi({"7100": "100"})
    draft = _draft_multi([("L1", "7100", "-50")])
    # The schema accepts the value; comparator skips it.
    assert compare_to_budget(draft, ctx) == []


# ---------- Multi-account / multi-line ----------

def test_two_postings_same_line_both_checked():
    """A line with two debit postings produces two BudgetCheck entries."""
    ctx = _ctx_multi({"7100": "1000", "7200": "1000"})
    draft = _draft_multi([
        ("L1", "7100", "500"),
        ("L1", "7200", "500"),
    ])
    results = compare_to_budget(draft, ctx)
    assert len(results) == 2
    statuses = {r.account_number: r.status for r in results}
    assert statuses["7100"] == BudgetStatus.WITHIN_BUDGET
    assert statuses["7200"] == BudgetStatus.WITHIN_BUDGET


def test_three_lines_three_statuses():
    """OVER + WARNING + WITHIN in one invoice."""
    ctx = _ctx_multi(
        {"7100": "100", "7200": "100", "7300": "1000"},
        ytd={"7100": "50", "7200": "85", "7300": "0"},
    )
    draft = _draft_multi([
        ("L1", "7100", "100"),  # 50+100=150 > 100 → OVER
        ("L2", "7200", "5"),    # 85+5=90 → 90% → WARNING
        ("L3", "7300", "10"),   # 1% → WITHIN
    ])
    results = compare_to_budget(draft, ctx)
    statuses = {r.line_id: r.status for r in results}
    assert statuses["L1"] == BudgetStatus.OVER_BUDGET
    assert statuses["L2"] == BudgetStatus.WARNING
    assert statuses["L3"] == BudgetStatus.WITHIN_BUDGET


def test_credit_only_then_debit_only_skips_credits():
    """Skip credit postings, check only debits."""
    ctx = _ctx_multi({"7100": "1000"})
    draft = DraftAllocations(
        invoice_number="INV",
        lines=[DraftLineAllocation(
            line_id="L1", description="d",
            postings=[
                Posting(account_number="7100", account_name="O", fund_id="GEN",
                        fund_name="G", debit_amount=Decimal("0"),
                        credit_amount=Decimal("100"),
                        restriction_class=RestrictionClass.WITHOUT_RESTRICTION),
                Posting(account_number="7100", account_name="O", fund_id="GEN",
                        fund_name="G", debit_amount=Decimal("100"),
                        credit_amount=Decimal("0"),
                        restriction_class=RestrictionClass.WITHOUT_RESTRICTION),
            ],
            total_debits=Decimal("100"), total_credits=Decimal("100"),
            balanced=True,
        )],
        document_total_debits=Decimal("100"),
        document_total_credits=Decimal("100"),
        document_balanced=True,
    )
    results = compare_to_budget(draft, ctx)
    assert len(results) == 1  # only the debit
    assert results[0].this_invoice == Decimal("100")


# ---------- Decimal precision ----------

def test_high_precision_decimals():
    """Comparator should not lose cents to floating-point conversion."""
    ctx = _ctx_multi({"7100": "1000.00"}, ytd={"7100": "999.99"})
    draft = _draft_multi([("L1", "7100", "0.01")])
    results = compare_to_budget(draft, ctx)
    # 999.99 + 0.01 = 1000.00 — exactly at 100% — > 80% threshold but not > annual
    assert results[0].status == BudgetStatus.WARNING
    assert results[0].after == Decimal("1000.00")
    assert results[0].remaining == Decimal("0.00")


def test_consumed_pct_for_zero_budget_clamped():
    """Zero budget with positive invoice gives 999.0 (sentinel for inf)."""
    ctx = _ctx_multi({"7100": "0"})
    draft = _draft_multi([("L1", "7100", "100")])
    results = compare_to_budget(draft, ctx)
    assert results[0].status == BudgetStatus.OVER_BUDGET
    assert results[0].consumed_pct == 999.0


# ---------- Spreadsheet parser edge cases ----------

def test_parser_empty_xlsx_returns_no_budget():
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"  # type: ignore[union-attr]
    buf = io.BytesIO()
    wb.save(buf)
    out = parse_spreadsheet(buf.getvalue(), "empty.xlsx")
    assert "budget" not in out


def test_parser_ignores_blank_rows():
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"  # type: ignore[union-attr]
    ws.append(["account_number", "annual_budget"])  # type: ignore[union-attr]
    ws.append([])  # type: ignore[union-attr]
    ws.append(["7100", 1000])  # type: ignore[union-attr]
    ws.append([None, None])  # type: ignore[union-attr]
    ws.append(["7200", 2000])  # type: ignore[union-attr]
    buf = io.BytesIO()
    wb.save(buf)
    out = parse_spreadsheet(buf.getvalue(), "b.xlsx")
    assert len(out["budget"]["accounts"]) == 2


def test_parser_handles_account_number_as_float():
    """Excel often loads numeric account numbers as floats (7100.0)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"  # type: ignore[union-attr]
    ws.append(["account_number", "annual_budget"])  # type: ignore[union-attr]
    ws.append([7100.0, 1000])  # type: ignore[union-attr]
    buf = io.BytesIO()
    wb.save(buf)
    out = parse_spreadsheet(buf.getvalue(), "b.xlsx")
    # Number should be normalized to "7100" (no trailing .0)
    assert "7100" in out["budget"]["accounts"]


def test_parser_currency_strings_in_amounts():
    """Amounts like $24000.00 with $ sign must parse as Decimal('24000.00')."""
    # CSV with quoted comma value
    csv_text = 'account_number,annual_budget\n7100,"$24,000.00"\n'
    out = parse_spreadsheet(csv_text.encode("utf-8"), "b.csv")
    assert out["budget"]["accounts"]["7100"]["annual_total"] == Decimal("24000.00")


def test_parser_explicit_zero_budget_preserved():
    """Explicit zero annual_budget is preserved (church may want to flag any spend)."""
    csv_text = "account_number,annual_budget\n7100,0\n7200,1000\n"
    out = parse_spreadsheet(csv_text.encode("utf-8"), "b.csv")
    accts = out["budget"]["accounts"]
    # Explicit zero is preserved (any positive invoice → OVER_BUDGET in comparator)
    assert "7100" in accts
    assert accts["7100"]["annual_total"] == Decimal("0")
    assert accts["7200"]["annual_total"] == Decimal("1000")


def test_parser_implicit_zero_budget_skipped():
    """Rows with no annual and no monthly data are skipped."""
    csv_text = "account_number,annual_budget,jan,feb\n7100,,,\n7200,1000,,\n"
    out = parse_spreadsheet(csv_text.encode("utf-8"), "b.csv")
    accts = out["budget"]["accounts"]
    assert "7100" not in accts
    assert "7200" in accts


# ---------- BudgetCheck schema invariants ----------

def test_budget_check_after_equals_ytd_plus_invoice():
    ctx = _ctx_multi({"7100": "1000"}, ytd={"7100": "300"})
    draft = _draft_multi([("L1", "7100", "200")])
    results = compare_to_budget(draft, ctx)
    bc = results[0]
    assert bc.after == bc.ytd_actual + bc.this_invoice
    assert bc.remaining == bc.annual_budget - bc.after


def test_budget_check_reason_has_expected_keywords():
    ctx_over = _ctx_multi({"7100": "100"}, ytd={"7100": "150"})
    draft = _draft_multi([("L1", "7100", "10")])
    over_result = compare_to_budget(draft, ctx_over)[0]
    assert "OVER BUDGET" in over_result.reason

    ctx_warn = _ctx_multi({"7100": "100"}, ytd={"7100": "75"})
    draft = _draft_multi([("L1", "7100", "10")])
    warn_result = compare_to_budget(draft, ctx_warn)[0]
    assert "WARNING" in warn_result.reason

    ctx_within = _ctx_multi({"7100": "1000"}, ytd={"7100": "0"})
    draft = _draft_multi([("L1", "7100", "10")])
    within_result = compare_to_budget(draft, ctx_within)[0]
    assert "Within budget" in within_result.reason
