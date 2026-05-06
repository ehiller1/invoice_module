"""Tests for budget extraction from spreadsheets."""
from __future__ import annotations
import io
from decimal import Decimal

import pytest
from openpyxl import Workbook

from backend.tools.spreadsheet_parser import parse_spreadsheet


def _make_xlsx(headers, rows, sheet_name="Budget"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_pure_annual_only_sheet():
    content = _make_xlsx(
        ["account_number", "account_name", "annual_budget"],
        [["7100", "Office", 24000], ["7200", "Utilities", 18000]],
    )
    out = parse_spreadsheet(content, "annual.xlsx")
    assert "budget" in out
    accts = out["budget"]["accounts"]
    assert accts["7100"]["annual_total"] == Decimal("24000")
    assert accts["7100"]["jan"] == Decimal("0")
    assert out["budget"]["annual_total"] == Decimal("42000")


def test_pure_monthly_sheet_derives_annual():
    content = _make_xlsx(
        ["account_number"] + ["jan", "feb", "mar", "apr", "may", "jun",
                              "jul", "aug", "sep", "oct", "nov", "dec"],
        [["7100"] + [1000] * 12],
    )
    out = parse_spreadsheet(content, "monthly.xlsx")
    accts = out["budget"]["accounts"]
    assert accts["7100"]["annual_total"] == Decimal("12000")
    assert accts["7100"]["jan"] == Decimal("1000")


def test_inconsistent_monthly_and_annual_warns():
    content = _make_xlsx(
        ["account_number", "annual_budget", "jan", "feb", "mar"],
        [["7100", 12000, 1000, 1000, 1000]],  # monthly sums to 3000, annual says 12000
    )
    out = parse_spreadsheet(content, "mixed.xlsx")
    assert any("disagrees" in w for w in out.get("warnings", []))
    # Explicit annual wins
    assert out["budget"]["accounts"]["7100"]["annual_total"] == Decimal("12000")


def test_consistent_monthly_and_annual_no_warning():
    content = _make_xlsx(
        ["account_number", "annual_budget"]
        + ["jan", "feb", "mar", "apr", "may", "jun",
           "jul", "aug", "sep", "oct", "nov", "dec"],
        [["7100", 12000] + [1000] * 12],
    )
    out = parse_spreadsheet(content, "consistent.xlsx")
    assert all("disagrees" not in w for w in out.get("warnings", []))


def test_csv_annual_only():
    csv_text = "account_number,account_name,annual_budget\n7100,Office,24000\n7200,Utilities,18000\n"
    out = parse_spreadsheet(csv_text.encode("utf-8"), "budget.csv")
    accts = out["budget"]["accounts"]
    assert accts["7100"]["annual_total"] == Decimal("24000")


def test_multi_sheet_xlsx_accounts_and_budget():
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws.append(["account_number", "annual_budget"])
    ws.append(["7100", 24000])
    ws2 = wb.create_sheet("Accounts")
    ws2.append(["account_number", "account_name", "account_type"])
    ws2.append(["7100", "Office Supplies", "Expense"])
    buf = io.BytesIO()
    wb.save(buf)
    out = parse_spreadsheet(buf.getvalue(), "combined.xlsx")
    assert len(out["accounts"]) == 1
    assert "7100" in out["budget"]["accounts"]


def test_unknown_columns_ignored():
    content = _make_xlsx(
        ["account_number", "annual_budget", "extra_col", "comment"],
        [["7100", 24000, "ignore", "this is a memo"]],
    )
    out = parse_spreadsheet(content, "x.xlsx")
    assert out["budget"]["accounts"]["7100"]["annual_total"] == Decimal("24000")
